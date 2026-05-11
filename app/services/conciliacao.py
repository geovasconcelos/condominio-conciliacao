"""
Serviço de análise e conciliação de cobranças condominiais.
Compara a planilha de dados (004A) com a planilha de parâmetros.
"""
import os
import numpy as np
import pandas as pd

from app.services.parametros import ler_parametros

# ── Helpers ────────────────────────────────────────────────────────────────────

def _br_to_float(s):
    if pd.isna(s):
        return np.nan
    s = str(s).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("(","").replace(")","").replace(".","").replace(",",".")
    try:
        return float(s) * (-1 if neg else 1)
    except ValueError:
        return np.nan


def _parse_date(s):
    if pd.isna(s):
        return pd.NaT
    try:
        return pd.to_datetime(str(s).strip(), dayfirst=True)
    except Exception:
        return pd.NaT


def _fmt_brl(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")


NUMERIC_COLS = [
    "Tarifa Liquidação Boleto","Taxa de Água","Medição e Leitura de Água",
    "Taxa Ordinária","Taxa Extra -  Modernização Elevadores",
    "Taxa Extra Manut. Piscina","Taxa Extra - Melhorias no Condomínio",
    "Taxa Extra - Aquisição de Gerador",
    "Taxa Extra  - Reforma - Aquisição - Equipamentos Academia",
    "Receita com Multas","Outros","Total",
]

# Mapeamento nome do parâmetro → coluna no 004A
EXTRA_COL_MAP = {
    "Taxa Extra -  Modernização Elevadores":              "Taxa Extra -  Modernização Elevadores",
    "Taxa Extra  - Reforma - Aquisição - Equipamentos Academia": "Taxa Extra  - Reforma - Aquisição - Equipamentos Academia",
    "Taxa Extra - Aquisição de Gerador":                  "Taxa Extra - Aquisição de Gerador",
    "Taxa Extra - Melhorias no Condomínio":               "Taxa Extra - Melhorias no Condomínio",
    "Taxa Extra Manut. Piscina":                          "Taxa Extra Manut. Piscina",
}


# ── Carregamento dos dados ──────────────────────────────────────────────────────

def _carregar_dados(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=3)
    df = df[df["Tipo Lançamento"].notna()].copy()

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = df[col].apply(_br_to_float)

    df["Vencimento_dt"] = df["Vencimento"].apply(_parse_date)
    df["Credito_dt"]    = df["Crédito"].apply(_parse_date)
    df["dias_atraso"]   = (df["Credito_dt"] - df["Vencimento_dt"]).dt.days
    df["Unidade"]       = df["Unidade/Bloco"].apply(
        lambda x: str(int(x)) if pd.notna(x) else ""
    )
    return df


# ── Análise principal ───────────────────────────────────────────────────────────

def processar_conciliacao(path_params: str, path_dados: str,
                          session_id: str, output_dir: str) -> dict:
    params = ler_parametros(path_params)
    df     = _carregar_dados(path_dados)

    df_normal = df[df["Tipo Cobrança"] == "NORMAL"].copy()
    df_extra  = df[df["Tipo Cobrança"] == "EXTRA"].copy()
    df_acordo = df[df["Tipo Cobrança"] == "ACORDO"].copy()

    unidades_dados   = set(df["Unidade"].unique()) - {""}
    unidades_params  = set(params["unidades"].keys())
    sindicos         = set(params["sindicos"])

    # ── 1. Campos faltantes no parâmetro ──────────────────────
    campos_faltantes = params["campos_faltantes"]

    # ── 2. Unidades sem parâmetro / parâmetro sem dados ───────
    sem_parametro = sorted(unidades_dados - unidades_params)
    sem_dados     = sorted(unidades_params - unidades_dados)

    # ── 3. Taxa Ordinária ──────────────────────────────────────
    inconsistencias_taxa = []
    taxa_padrao = params["taxa_ord_padrao"] or 0.0

    for unidade, grp in df_normal.groupby("Unidade"):
        p = params["unidades"].get(unidade)
        if not p:
            continue

        isento = unidade in sindicos and params["isencao_sindico"]
        taxa_esperada = 0.0 if isento else (p["taxa_ordinaria"] or taxa_padrao)

        # Agrupa por mês e verifica apenas o maior valor do mês.
        # Cada mês pode ter múltiplas linhas NORMAL (taxa ordinária + outros encargos);
        # a linha de outros encargos tem taxa=0 e não deve ser verificada individualmente.
        grp = grp.copy()
        grp["_mes"] = grp["Vencimento_dt"].dt.to_period("M")
        for mes, grp_mes in grp.groupby("_mes"):
            taxa_real = grp_mes["Taxa Ordinária"].max()
            if pd.isna(taxa_real):
                taxa_real = 0.0
            diff = abs(taxa_real - taxa_esperada)
            if diff > 0.05:
                row = grp_mes.iloc[0]
                inconsistencias_taxa.append({
                    "Unidade":         unidade,
                    "Vencimento":      row["Vencimento_dt"].strftime("%m/%Y") if pd.notna(row["Vencimento_dt"]) else "-",
                    "Esperado (R$)":   taxa_esperada,
                    "Encontrado (R$)": taxa_real,
                    "Diferença (R$)":  taxa_real - taxa_esperada,
                    "Motivo":          "Síndico isento" if isento else "Divergência de taxa",
                })

    # ── 3b. Inadimplência — boletos ausentes na 004A ─────────────
    # A 004A só contém registros PAGOS. Meses sem registro = inadimplência.
    # Exclusão: unidades com ACORDO posterior ao mês ausente tiveram a dívida regularizada.
    boletos_ausentes = []

    all_months = pd.period_range(
        df["Vencimento_dt"].min().to_period("M"),
        df["Vencimento_dt"].max().to_period("M"),
        freq="M"
    )
    df_n_mes = df_normal.copy()
    df_n_mes["_mes"] = df_n_mes["Vencimento_dt"].dt.to_period("M")
    df_e_mes = df_extra.copy()
    df_e_mes["_mes"] = df_e_mes["Vencimento_dt"].dt.to_period("M")
    df_ac = df_acordo.copy()
    df_ac["_mes"] = df_ac["Vencimento_dt"].dt.to_period("M")
    acordo_meses = {
        u: set(g["_mes"].unique())
        for u, g in df_ac.groupby("Unidade")
    }

    # NORMAL: boleto de taxa ordinária ausente
    if taxa_padrao > 0:
        for unidade, p in params["unidades"].items():
            isento = unidade in sindicos and params["isencao_sindico"]
            if isento:
                continue
            taxa_esp = p["taxa_ordinaria"] or taxa_padrao
            if not taxa_esp or taxa_esp < 0.05:
                continue
            regs_u   = df_n_mes[df_n_mes["Unidade"] == unidade]
            u_acords = acordo_meses.get(unidade, set())
            for mes in all_months:
                regs_mes = regs_u[regs_u["_mes"] == mes]
                max_taxa = float(regs_mes["Taxa Ordinária"].max()) if len(regs_mes) > 0 else 0.0
                if pd.isna(max_taxa):
                    max_taxa = 0.0
                if max_taxa < 0.05:
                    if any(a >= mes for a in u_acords):
                        continue  # débito regularizado via ACORDO
                    boletos_ausentes.append({
                        "Unidade":     unidade,
                        "Competência": f"{mes.month:02d}/{mes.year}",
                        "tipo_inadin": "NORMAL",
                    })

    # EXTRA: boleto de taxas extras ausente (unidades com tem_taxa_extra=True)
    if params["taxas_extras"]:
        def _extra_ativo(mes):
            for te in params["taxas_extras"]:
                if not te["valor"] or te["valor"] < 0.05:
                    continue
                if te["inicio"] and te["fim"]:
                    p_ini = pd.Period(year=te["inicio"][1], month=te["inicio"][0], freq="M")
                    p_fim = pd.Period(year=te["fim"][1],    month=te["fim"][0],    freq="M")
                    if p_ini <= mes <= p_fim:
                        return True
                else:
                    return True
            return False

        meses_extra_ativos = [m for m in all_months if _extra_ativo(m)]
        normal_set = {(b["Unidade"], b["Competência"]) for b in boletos_ausentes}

        for unidade, p in params["unidades"].items():
            if not p.get("tem_taxa_extra"):
                continue
            regs_u   = df_e_mes[df_e_mes["Unidade"] == unidade]
            u_acords = acordo_meses.get(unidade, set())
            for mes in meses_extra_ativos:
                comp_str = f"{mes.month:02d}/{mes.year}"
                if (unidade, comp_str) in normal_set:
                    continue  # já listado como NORMAL (inclui encargos extras)
                if len(regs_u[regs_u["_mes"] == mes]) > 0:
                    continue  # boleto EXTRA existe
                if any(a >= mes for a in u_acords):
                    continue  # regularizado via ACORDO
                boletos_ausentes.append({
                    "Unidade":     unidade,
                    "Competência": comp_str,
                    "tipo_inadin": "EXTRA",
                })

    # ── 4. Taxa de Água ────────────────────────────────────────
    problemas_agua = []
    if "Taxa de Água" in df.columns:
        for unidade, grp in df_normal.groupby("Unidade"):
            meses_sem_agua = grp[
                grp["Taxa de Água"].isna() | (grp["Taxa de Água"] == 0)
            ]
            for _, row in meses_sem_agua.iterrows():
                problemas_agua.append({
                    "Unidade":    unidade,
                    "Vencimento": row["Vencimento_dt"].strftime("%d/%m/%Y") if pd.notna(row["Vencimento_dt"]) else "-",
                    "Problema":   "Taxa de Água ausente ou zero",
                    "Valor":      row["Taxa de Água"],
                })

    # ── 5. Medição e Leitura de Água ───────────────────────────
    problemas_medicao = []
    ref_medicao = params["taxa_medicao"] or 0.0
    if "Medição e Leitura de Água" in df.columns and ref_medicao > 0:
        for unidade, grp in df_normal.groupby("Unidade"):
            for _, row in grp.iterrows():
                med = row["Medição e Leitura de Água"]
                if pd.isna(med) or med == 0:
                    problemas_medicao.append({
                        "Unidade":    unidade,
                        "Vencimento": row["Vencimento_dt"].strftime("%d/%m/%Y") if pd.notna(row["Vencimento_dt"]) else "-",
                        "Problema":   "Medição ausente ou zero",
                        "Esperado":   ref_medicao,
                        "Encontrado": med,
                    })
                elif abs(med - ref_medicao) > 0.05:
                    problemas_medicao.append({
                        "Unidade":    unidade,
                        "Vencimento": row["Vencimento_dt"].strftime("%d/%m/%Y") if pd.notna(row["Vencimento_dt"]) else "-",
                        "Problema":   "Valor divergente",
                        "Esperado":   ref_medicao,
                        "Encontrado": med,
                    })

    # ── 6. Taxas Extras ────────────────────────────────────────
    inconsistencias_extra = []
    for taxa_param in params["taxas_extras"]:
        nome    = taxa_param["nome"]
        col     = EXTRA_COL_MAP.get(nome)
        valor_p = taxa_param["valor"]
        inicio  = taxa_param["inicio"]
        fim     = taxa_param["fim"]

        if not col or col not in df.columns:
            continue

        df_e = df_extra[df_extra["Unidade"].isin(
            [u for u, p in params["unidades"].items() if p["tem_taxa_extra"]]
        )].copy()

        for unidade, grp in df_e.groupby("Unidade"):
            for _, row in grp.iterrows():
                venc = row["Vencimento_dt"]
                if pd.isna(venc):
                    continue

                # Verifica se está no período esperado (comparação por data, não tupla)
                no_periodo = True
                if inicio and fim:
                    inicio_date = pd.Timestamp(year=inicio[1], month=inicio[0], day=1)
                    fim_date    = pd.Timestamp(year=fim[1],    month=fim[0],    day=28)
                    no_periodo  = inicio_date <= venc <= fim_date

                valor_real = row[col]
                if pd.isna(valor_real):
                    valor_real = 0.0

                if no_periodo and abs(valor_real - valor_p) > 0.05:
                    inconsistencias_extra.append({
                        "Taxa":          nome,
                        "Unidade":       unidade,
                        "Vencimento":    venc.strftime("%d/%m/%Y"),
                        "Esperado (R$)": valor_p,
                        "Encontrado (R$)": valor_real,
                        "Diferença (R$)": valor_real - valor_p,
                    })
                elif not no_periodo and valor_real > 0:
                    inconsistencias_extra.append({
                        "Taxa":          nome,
                        "Unidade":       unidade,
                        "Vencimento":    venc.strftime("%d/%m/%Y"),
                        "Esperado (R$)": 0.0,
                        "Encontrado (R$)": valor_real,
                        "Diferença (R$)": valor_real,
                        "Motivo":        "Cobrança fora do período vigente",
                    })

    # ── 7. Multa ────────────────────────────────────────────────
    carencia    = params["carencia_dias"]
    pct_multa_p = params["pct_multa"]

    # Síndico isento não entra nas verificações de atraso e multa
    if params["isencao_sindico"] and sindicos:
        df_multa = df_normal[~df_normal["Unidade"].isin(sindicos)].copy()
    else:
        df_multa = df_normal.copy()

    # Inadimplente = pago com atraso OU boleto sem crédito com vencimento já passado
    hoje = pd.Timestamp.today().normalize()
    sem_credito = (
        df_multa["Credito_dt"].isna() &
        df_multa["Vencimento_dt"].notna() &
        (df_multa["Vencimento_dt"] < hoje)
    )
    df_multa["atrasado_real"] = (df_multa["dias_atraso"] > carencia) | sem_credito
    df_multa["tem_multa"]     = df_multa["Receita com Multas"].fillna(0) > 0

    atrasados_reais     = int(df_multa["atrasado_real"].sum())
    atrasados_sem_multa = df_multa[df_multa["atrasado_real"] & ~df_multa["tem_multa"]]
    total_criticos      = len(atrasados_sem_multa)

    com_multa_df = df_multa[df_multa["atrasado_real"] & df_multa["tem_multa"]].copy()
    com_multa_df = com_multa_df[com_multa_df["Taxa Ordinária"].fillna(0) > 0].copy()
    com_multa_df["pct_multa_calc"] = (
        com_multa_df["Receita com Multas"] / com_multa_df["Taxa Ordinária"] * 100
    )
    multa_inconsistente = com_multa_df[
        (com_multa_df["pct_multa_calc"] < pct_multa_p * 0.75) |
        (com_multa_df["pct_multa_calc"] > pct_multa_p * 1.5)
    ]
    multa_em_zero = df_multa[
        (df_multa["Taxa Ordinária"].fillna(0) == 0) & (df_multa["Receita com Multas"].fillna(0) > 0)
    ]

    # ── 8. Estatísticas gerais ────────────────────────────────
    total_registros = len(df)
    total_unidades  = df["Unidade"].nunique()
    periodo_inicio  = df["Vencimento_dt"].min()
    periodo_fim     = df["Vencimento_dt"].max()
    total_emissao   = df["Total"].sum()

    no_prazo            = int((df_normal["dias_atraso"] <= 0).sum())
    compensacao_banc    = int(((df_normal["dias_atraso"] > 0) & (df_normal["dias_atraso"] <= carencia)).sum())

    bins   = [0, 1, 2, 4, 7, 15, 30, 9999]
    labels = ["1 dia","2 dias","3-4 dias","5-7 dias","8-15 dias","16-30 dias",">30 dias"]
    dist_atraso = (
        pd.cut(df_normal[df_normal["dias_atraso"] > 0]["dias_atraso"],
               bins=bins, labels=labels, right=True)
        .value_counts().sort_index().to_dict()
    )

    multa_media = round(com_multa_df["pct_multa_calc"].mean(), 2) if len(com_multa_df) else 0.0
    multa_min   = round(com_multa_df["pct_multa_calc"].min(), 2)  if len(com_multa_df) else 0.0
    multa_max   = round(com_multa_df["pct_multa_calc"].max(), 2)  if len(com_multa_df) else 0.0

    # Outliers
    q1  = df["Total"].quantile(0.25)
    q3  = df["Total"].quantile(0.75)
    iqr = q3 - q1
    outliers_n = len(df[
        ((df["Total"] < q1 - 1.5 * iqr) | (df["Total"] > q3 + 1.5 * iqr)) &
        (df["Tipo Cobrança"] == "NORMAL")
    ])

    resultado = {
        # Geral
        "total_registros":    total_registros,
        "total_unidades":     total_unidades,
        "periodo_inicio":     periodo_inicio.strftime("%d/%m/%Y") if pd.notna(periodo_inicio) else "-",
        "periodo_fim":        periodo_fim.strftime("%d/%m/%Y")    if pd.notna(periodo_fim)    else "-",
        "total_emissao":      _fmt_brl(total_emissao),
        "qt_normal":          len(df_normal),
        "qt_extra":           len(df_extra),
        "qt_acordo":          len(df_acordo),
        # Parâmetros
        "taxa_padrao_param":  _fmt_brl(params["taxa_ord_padrao"]),
        "carencia_param":     carencia,
        "pct_multa_param":    pct_multa_p,
        "sindicos":           params["sindicos"],
        "isencao_sindico":    params["isencao_sindico"],
        "ref_medicao":        _fmt_brl(params["taxa_medicao"]),
        "qt_taxas_extras_param": len(params["taxas_extras"]),
        # Validação de parâmetros
        "campos_faltantes":   campos_faltantes,
        "sem_parametro":      sem_parametro,
        "sem_dados":          sem_dados,
        # Taxa ordinária
        "qt_inconsist_taxa":  len(inconsistencias_taxa),
        # Inadimplência (boletos ausentes)
        "qt_boletos_ausentes": len(boletos_ausentes),
        "unidades_inadimplentes": sorted(set(b["Unidade"] for b in boletos_ausentes)),
        # Água
        "qt_prob_agua":       len(problemas_agua),
        "qt_prob_medicao":    len(problemas_medicao),
        # Taxa extra
        "qt_inconsist_extra": len(inconsistencias_extra),
        # Atrasos
        "no_prazo":           no_prazo,
        "compensacao":        compensacao_banc,
        "atrasados_reais":    atrasados_reais,
        "total_criticos":     total_criticos,
        "dist_atraso":        {str(k): int(v) for k, v in dist_atraso.items()},
        # Multa
        "total_com_multa":    len(com_multa_df),
        "multa_media":        multa_media,
        "multa_min":          multa_min,
        "multa_max":          multa_max,
        "multa_inconsistente_qt": len(multa_inconsistente),
        "multa_em_zero_qt":   len(multa_em_zero),
        # Atípicos e acordos
        "total_outliers":     outliers_n,
        "total_acordos":      len(df_acordo),
        "unidades_acordo":    df_acordo["Unidade"].unique().tolist(),
        "valor_acordos":      _fmt_brl(df_acordo["Total"].sum()),
    }

    # Gera Excel
    _gerar_excel(df, atrasados_sem_multa,
                 multa_inconsistente, multa_em_zero,
                 pd.DataFrame(inconsistencias_taxa),
                 pd.DataFrame(problemas_agua),
                 pd.DataFrame(problemas_medicao),
                 pd.DataFrame(inconsistencias_extra),
                 pd.DataFrame(boletos_ausentes),
                 sem_parametro, sem_dados, campos_faltantes,
                 resultado, params, output_dir, session_id)

    return resultado


# ── Geração do Excel ────────────────────────────────────────────────────────────

def _add_nota(ws, row, ncols, text):
    """Linha de cabeçalho explicativo padronizada para todas as abas."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.fill  = PatternFill("solid", fgColor="E3F2FD")
    c.font  = Font(name="Calibri", size=9, color="1A3C6E")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    s = Side(style="medium", color="1E88E5")
    c.border = Border(left=s)
    ws.row_dimensions[row].height = 72


def _aba_inadimplencia(ws, df, params, df_inadimplentes):
    """Aba 1 — Relatório de inadimplência detalhado por unidade."""
    from collections import defaultdict
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    RED    = "B71C1C"; RLIGHT = "FFEBEE"
    NAVY   = "1A3C6E"; WHITE  = "FFFFFF"
    LGRAY  = "F0F4F8"; GOLD   = "FFF8E1"
    MGRAY  = "ECEFF1"

    def _f(c):   return PatternFill("solid", fgColor=c)
    def _ft(bold=False, color="1A2B3C", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def _bd(c="B0BEC5"):
        s = Side(style="thin", color=c); return Border(left=s, right=s, top=s, bottom=s)
    def _al(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=False)

    NCOLS = 9
    COL_W = [13, 12, 8, 42, 14, 14, 14, 13, 14]
    COL_H = ["Vencimento","Competência","Cód.","Descrição",
             "Valor (R$)","Juros (R$)","Multa (R$)","Atualiz. (R$)","Total (R$)"]
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Título ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws["A1"]
    c.value = "RELATÓRIO DE INADIMPLÊNCIA"
    c.font  = _ft(bold=True, color=WHITE, size=14)
    c.fill  = _f(RED); c.alignment = _al("center")

    _add_nota(ws, 2, NCOLS,
        "O QUE ESTA ABA VERIFICA: dois tipos de inadimplência são detectados — "
        "(A) BOLETO NORMAL ausente: unidade sem registro NORMAL com taxa ordinária > R$ 0,05 no mês; "
        "a composição exibe taxa ordinária, água, medição e taxas extras (quando aplicável). "
        "(B) BOLETO EXTRA ausente: unidades com taxa extra no cadastro que não possuem registro EXTRA "
        "na 004A em meses em que pelo menos uma taxa extra está vigente, mas que pagaram o boleto NORMAL; "
        "a composição exibe apenas as taxas extras devidas. "
        "A base 004A contém apenas registros de pagamentos efetivados; ausência de registro = ausência de pagamento.\n"
        "CRITÉRIOS DE EXCLUSÃO: (1) Síndico isento — unidade não é verificada quando isenção está configurada; "
        "(2) Regularização via ACORDO — se a unidade possui registro de ACORDO na 004A com vencimento igual ou "
        "posterior ao mês ausente, o débito é considerado regularizado e o mês não é listado.\n"
        "CÁLCULO DE ENCARGOS: Multa = " + f"{params['pct_multa']:.2f}% sobre o valor principal | "
        f"Juros = {params['pct_juros']:.2f}% a.m. pró-rata (dias corridos após carência de "
        f"{params.get('carencia_dias', 4)} dias) | (*) Taxa de Água estimada pelo último boleto pago da unidade."
    )

    ws.row_dimensions[3].height = 15
    ws.merge_cells(f"A3:{get_column_letter(NCOLS)}3")
    c = ws["A3"]
    hoje_str = pd.Timestamp.today().strftime("%d/%m/%Y")
    c.value = (f"Multa: {params['pct_multa']:.2f}%  |  "
               f"Juros: {params['pct_juros']:.2f}% a.m. (pró-rata)  |  "
               f"Valores atualizados até {hoje_str}  |  "
               f"(*) Taxa de Água estimada a partir do último boleto pago da unidade")
    c.font  = _ft(size=9, color="455A64"); c.fill = _f("FAFAFA")
    c.alignment = _al("left")

    if df_inadimplentes is None or len(df_inadimplentes) == 0:
        ws.merge_cells(f"A4:{get_column_letter(NCOLS)}4")
        ws["A4"].value = "Nenhuma inadimplência detectada."
        ws["A4"].font  = _ft(bold=True, color="2E7D32", size=11)
        return

    r = 4

    # Parâmetros de cálculo
    pct_multa  = float(params.get("pct_multa")  or 2.0)
    pct_juros  = float(params.get("pct_juros")  or 1.0)
    carencia   = int(params.get("carencia_dias") or 4)
    dia_venc   = int(params.get("dia_vencimento") or 10)
    taxa_med_p = float(params.get("taxa_medicao") or 0.0)
    taxas_ext  = params.get("taxas_extras", [])
    ref_date   = pd.Timestamp.today().normalize()

    df_normal = df[df["Tipo Cobrança"] == "NORMAL"].copy() if "Tipo Cobrança" in df.columns else df.copy()

    # Último boleto pago por unidade → estima água e medição
    agua_est = {}
    med_est  = {}
    for unidade, grp in df_normal.groupby("Unidade"):
        col_a = "Taxa de Água"
        col_m = "Medição e Leitura de Água"
        grp_a = grp[grp[col_a].notna() & (grp[col_a] > 0)] if col_a in grp.columns else pd.DataFrame()
        if len(grp_a):
            agua_est[unidade] = float(grp_a.loc[grp_a["Vencimento_dt"].idxmax(), col_a])
        grp_m = grp[grp[col_m].notna() & (grp[col_m] > 0)] if col_m in grp.columns else pd.DataFrame()
        if len(grp_m):
            med_est[unidade]  = float(grp_m.loc[grp_m["Vencimento_dt"].idxmax(), col_m])

    # Agrupa meses ausentes por unidade: (competência, tipo_inadin)
    unidades_meses = defaultdict(list)
    for _, row in df_inadimplentes.iterrows():
        tipo = str(row["tipo_inadin"]) if "tipo_inadin" in row.index else "NORMAL"
        unidades_meses[str(row["Unidade"])].append((str(row["Competência"]), tipo))

    cat_totals  = defaultdict(lambda: {"valor": 0.0, "juros": 0.0, "multa": 0.0, "total": 0.0})
    grand       = {"valor": 0.0, "juros": 0.0, "multa": 0.0, "total": 0.0}

    def _col_hdr(r):
        for ci, h in enumerate(COL_H, 1):
            c = ws.cell(row=r, column=ci)
            c.value = h
            c.font  = _ft(bold=True, color=WHITE, size=10)
            c.fill  = _f(NAVY); c.border = _bd(); c.alignment = _al("center")
        ws.row_dimensions[r].height = 18

    def _charge_row(r, venc_s, comp_s, desc, valor, juros, multa, bg=WHITE):
        atualiz = 0.0
        total   = round(valor + juros + multa, 2)
        row_vals = [venc_s, comp_s, "-", desc,
                    round(valor,2), round(juros,2), round(multa,2), atualiz, total]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci)
            c.value = v; c.fill = _f(bg); c.border = _bd()
            c.font  = _ft(size=10)
            if ci >= 5:
                c.alignment = _al("right")
                if isinstance(v, float): c.number_format = '#,##0.00'
            else:
                c.alignment = _al("left")
        ws.row_dimensions[r].height = 16
        return total

    for unidade in sorted(unidades_meses.keys(), key=lambda x: x.zfill(10)):
        competencias = sorted(unidades_meses[unidade])

        # Cabeçalho da unidade
        ws.merge_cells(f"A{r}:{get_column_letter(NCOLS)}{r}")
        c = ws[f"A{r}"]
        c.value = f"  UNIDADE  {unidade}"
        c.font  = _ft(bold=True, color=WHITE, size=11)
        c.fill  = _f(RED); c.alignment = _al("left")
        ws.row_dimensions[r].height = 22
        r += 1

        _col_hdr(r); r += 1

        u_tot = {"valor": 0.0, "juros": 0.0, "multa": 0.0, "total": 0.0}
        bg_toggle = 0

        for comp, tipo in sorted(competencias, key=lambda x: x[0]):
            mes, ano = int(comp[:2]), int(comp[3:])
            try:
                venc_date = pd.Timestamp(year=ano, month=mes, day=dia_venc)
            except Exception:
                venc_date = pd.Timestamp(year=ano, month=mes, day=1)
            venc_s = venc_date.strftime("%d/%m/%Y")
            dias   = max(0, (ref_date - venc_date).days - carencia)

            p_u    = params["unidades"].get(unidade, {})
            t_ord  = float(p_u.get("taxa_ordinaria") or params["taxa_ord_padrao"] or 0)
            t_agua = agua_est.get(unidade, 0.0)
            t_med  = med_est.get(unidade, taxa_med_p)

            charges = []
            # Boleto NORMAL ausente: todos os encargos; EXTRA ausente: só taxas extras
            if tipo == "NORMAL":
                if t_ord  > 0.05: charges.append(("Taxa Ordinária",            t_ord,  "Taxa Ordinária"))
                if t_agua > 0.05: charges.append(("Taxa de Água (*)",          t_agua, "Taxa de Água"))
                if t_med  > 0.05: charges.append(("Medição e Leitura de Água", t_med,  "Medição e Leitura de Água"))

            if p_u.get("tem_taxa_extra"):
                for te in taxas_ext:
                    if not te["valor"] or te["valor"] < 0.05:
                        continue
                    in_p = True
                    if te["inicio"] and te["fim"]:
                        d0 = pd.Timestamp(year=te["inicio"][1], month=te["inicio"][0], day=1)
                        d1 = pd.Timestamp(year=te["fim"][1],    month=te["fim"][0],    day=28)
                        in_p = d0 <= venc_date <= d1
                    if in_p:
                        charges.append((te["nome"], float(te["valor"]), te["nome"]))

            for desc, valor, cat_key in charges:
                multa  = valor * pct_multa / 100
                juros  = valor * pct_juros / 100 * (dias / 30) if dias > 0 else 0.0
                bg     = LGRAY if bg_toggle % 2 == 0 else WHITE
                total  = _charge_row(r, venc_s, comp, desc,
                                     valor, juros, multa, bg)
                r += 1; bg_toggle += 1

                for k, v in [("valor",valor),("juros",juros),("multa",multa),("total",total)]:
                    u_tot[k]              += v
                    cat_totals[cat_key][k] += v
                    grand[k]               += v

        # Linha de total da unidade
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = "Total"; c.font = _ft(bold=True, size=10)
        c.fill = _f(GOLD); c.border = _bd(); c.alignment = _al("right")
        for ci, key in zip([5,6,7], ["valor","juros","multa"]):
            c = ws.cell(row=r, column=ci)
            c.value = round(u_tot[key], 2); c.fill = _f(GOLD)
            c.font = _ft(bold=True, size=10); c.border = _bd()
            c.alignment = _al("right"); c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=8)                               # atualiz
        c.value = 0.0; c.fill = _f(GOLD); c.border = _bd()
        c.alignment = _al("right"); c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=9)                               # total
        c.value = round(u_tot["total"], 2); c.fill = _f(GOLD)
        c.font = _ft(bold=True, size=10); c.border = _bd()
        c.alignment = _al("right"); c.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 17
        r += 2  # linha em branco entre unidades

    # ── Resumo por categoria ────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:{get_column_letter(NCOLS)}{r}")
    c = ws[f"A{r}"]
    c.value = "Resumo por categoria"
    c.font  = _ft(bold=True, color=WHITE, size=12)
    c.fill  = _f(NAVY); c.alignment = _al("left")
    ws.row_dimensions[r].height = 24
    r += 1

    # Cabeçalhos do resumo: Descrição(A-D) | Valor(E) | Juros(F) | Multa(G) | Atualiz(H) | Total(I)
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "Descrição"; c.font = _ft(bold=True, color=WHITE, size=10)
    c.fill = _f(NAVY); c.border = _bd(); c.alignment = _al("left")
    for col_idx, h in zip([5,6,7,8,9],
                          ["Valor (R$)","Juros (R$)","Multa (R$)","Atualiz. (R$)","Total (R$)"]):
        c = ws.cell(row=r, column=col_idx)
        c.value = h; c.font = _ft(bold=True, color=WHITE, size=10)
        c.fill = _f(NAVY); c.border = _bd(); c.alignment = _al("center")
    ws.row_dimensions[r].height = 18
    r += 1

    cat_order = ["Taxa Ordinária","Taxa de Água","Medição e Leitura de Água"]
    extra_cats = [k for k in cat_totals if k not in cat_order]
    for i, cat in enumerate(cat_order + sorted(extra_cats)):
        if cat not in cat_totals:
            continue
        tots = cat_totals[cat]
        bg   = LGRAY if i % 2 == 0 else WHITE
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = cat.replace(" (*)", "")
        c.fill = _f(bg); c.font = _ft(size=10)
        c.border = _bd(); c.alignment = _al("left")
        for col_idx, (key, is_zero) in zip([5,6,7,8,9], [
                ("valor",False),("juros",False),("multa",False),("_",True),("total",False)]):
            c = ws.cell(row=r, column=col_idx)
            v = 0.0 if is_zero else tots.get(key, 0.0)
            c.value = round(v, 2); c.fill = _f(bg); c.font = _ft(size=10)
            c.border = _bd(); c.alignment = _al("right"); c.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 16
        r += 1

    # Total geral do resumo
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = "Total"; c.font = _ft(bold=True, size=10)
    c.fill = _f(GOLD); c.border = _bd(); c.alignment = _al("right")
    for col_idx, (key, is_zero) in zip([5,6,7,8,9], [
            ("valor",False),("juros",False),("multa",False),("_",True),("total",False)]):
        c = ws.cell(row=r, column=col_idx)
        v = 0.0 if is_zero else grand.get(key, 0.0)
        c.value = round(v, 2); c.fill = _f(GOLD)
        c.font = _ft(bold=True, size=10); c.border = _bd()
        c.alignment = _al("right"); c.number_format = '#,##0.00'
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"


def _aba_fluxo_pagamentos(ws, df, params):
    """Aba — Fluxo de pagamentos mensal por tipo de cobrança."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY  = "1A3C6E"; BLUE  = "1E88E5"; WHITE = "FFFFFF"
    LGRAY = "F0F4F8"; GOLD  = "FFF8E1"
    C_RED = "C62828"; C_GRN = "1B5E20"

    def _f(c): return PatternFill("solid", fgColor=c)
    def _ft(bold=False, color="1A2B3C", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def _bd():
        s = Side(style="thin", color="B0BEC5")
        return Border(left=s, right=s, top=s, bottom=s)
    def _al(h="center"):
        return Alignment(horizontal=h, vertical="center")

    carencia  = int(params.get("carencia_dias") or 4)
    sindicos  = set(params.get("sindicos") or [])
    isencao   = params.get("isencao_sindico", False)
    all_units = set(params["unidades"].keys())
    units_ext = {u for u, p in params["unidades"].items() if p["tem_taxa_extra"]}
    units_ord = (all_units - sindicos) if isencao else all_units

    df_n = df[df["Tipo Cobrança"] == "NORMAL"].copy()
    df_e = df[df["Tipo Cobrança"] == "EXTRA"].copy()
    df_a = df[df["Tipo Cobrança"] == "ACORDO"].copy()
    for _d in (df_n, df_e, df_a):
        _d["_mes"] = _d["Vencimento_dt"].dt.to_period("M")

    # Step 1 — all periods in ascending order
    all_periods = sorted(pd.unique(df["Vencimento_dt"].dropna().dt.to_period("M")))

    # Base columns for pro-rata multa allocation
    VALOR_COLS_BASE = [
        "Taxa Ordinária", "Taxa de Água", "Medição e Leitura de Água",
        "Taxa Extra -  Modernização Elevadores",
        "Taxa Extra  - Reforma - Aquisição - Equipamentos Academia",
        "Taxa Extra - Aquisição de Gerador",
        "Taxa Extra - Melhorias no Condomínio",
        "Taxa Extra Manut. Piscina",
    ]

    # Step 3 — payment reference list
    pmts = [
        {"nome": "Taxa Ordinária",            "col": "Taxa Ordinária",
         "df_src": df_n, "n_esp": len(units_ord), "units_set": units_ord,
         "p_ini": None, "p_fim": None, "add_outros": True},
        {"nome": "Taxa de Água",              "col": "Taxa de Água",
         "df_src": df_n, "n_esp": len(all_units), "units_set": all_units,
         "p_ini": None, "p_fim": None, "add_outros": False},
        {"nome": "Medição e Leitura de Água", "col": "Medição e Leitura de Água",
         "df_src": df_n, "n_esp": len(all_units), "units_set": all_units,
         "p_ini": None, "p_fim": None, "add_outros": False},
    ]
    for te in params.get("taxas_extras", []):
        col = EXTRA_COL_MAP.get(te["nome"])
        if not col or col not in df.columns:
            continue
        p_ini = pd.Period(year=te["inicio"][1], month=te["inicio"][0], freq="M") if te["inicio"] else None
        p_fim = pd.Period(year=te["fim"][1],    month=te["fim"][0],    freq="M") if te["fim"] else None
        pmts.append({
            "nome": te["nome"], "col": col,
            "df_src": df_e, "n_esp": len(units_ext), "units_set": units_ext,
            "p_ini": p_ini, "p_fim": p_fim, "add_outros": False,
        })

    HDRS = [
        "Competência", "Pagamentos do mês", "Valor esperado",
        "Unid. no vencimento", "Unid. com atraso",
        "Valor base pago", "Multa e juros pago", "Outros pagos",
        "Total pago", "Diferença de fluxo", "Unid. inadimplentes",
        "Lista de inadimplentes",
    ]
    NCOLS = len(HDRS)
    COL_W = [14, 17, 15, 19, 16, 16, 17, 14, 13, 17, 18, 40]
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws["A1"]
    c.value = "FLUXO DE PAGAMENTOS"
    c.font = _ft(bold=True, color=WHITE, size=13); c.fill = _f(NAVY); c.alignment = _al()

    _add_nota(ws, 2, NCOLS,
        "O QUE ESTA ABA VERIFICA: reconciliação mensal do fluxo de caixa por tipo de cobrança "
        "(Taxa Ordinária, Taxa de Água, Medição e Taxas Extras). Cada seção exibe uma linha por competência.\n"
        "COLUNAS PRINCIPAIS: "
        "Valor esperado = valor médio pago pelas unidades no prazo × total de unidades obrigadas "
        "(fallback: valor dos parâmetros quando não há pagadores no prazo); "
        f"Unid. no vencimento / com atraso = crédito dentro / fora da carência de {carencia} dias; "
        "Multa e juros pago = rateado pro-rata pela participação de cada cobrança no total do boleto; "
        "Outros pagos (somente Taxa Ordinária) = coluna 'Outros' da 004A + totais de ACORDO do período; "
        "Diferença de fluxo = Total pago − Valor esperado (vermelho = déficit, verde = superávit).\n"
        "ATENÇÃO: 'Unid. inadimplentes' e 'Lista de inadimplentes' refletem unidades sem registro "
        "desse tipo de cobrança na 004A para o período — podem incluir lacunas no arquivo-fonte "
        "que não representam inadimplência real. Para o relatório oficial de inadimplência, consulte a aba 'Inadimplência'."
    )

    r = 3

    for pmt in pmts:
        nome   = pmt["nome"]; col    = pmt["col"]
        df_src = pmt["df_src"]; n_esp  = pmt["n_esp"]
        units_s = pmt["units_set"]; p_ini  = pmt["p_ini"]; p_fim  = pmt["p_fim"]
        add_o  = pmt["add_outros"]

        if col not in df.columns:
            continue

        # Section header
        ws.merge_cells(f"A{r}:{get_column_letter(NCOLS)}{r}")
        c = ws[f"A{r}"]
        c.value = nome; c.font = _ft(bold=True, color=WHITE, size=11)
        c.fill = _f(BLUE); c.alignment = _al("left")
        ws.row_dimensions[r].height = 20; r += 1

        # Column headers
        for ci, h in enumerate(HDRS, 1):
            c = ws.cell(row=r, column=ci)
            c.value = h; c.font = _ft(bold=True, color=WHITE, size=9)
            c.fill = _f("2C5282"); c.border = _bd()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 28; r += 1

        # Relevant periods for this payment (respects inicio/fim for extras)
        rel_periods = [
            p for p in all_periods
            if (p_ini is None or p >= p_ini) and (p_fim is None or p <= p_fim)
        ]

        tot = {"pag": 0, "v_esp": 0.0, "no_v": 0, "atr": 0,
               "base": 0.0, "mj": 0.0, "out": 0.0, "tot": 0.0, "inad": 0}

        for idx, period in enumerate(rel_periods):
            comp_str = f"{period.month:02d}/{period.year}"
            df_per = df_src[df_src["_mes"] == period].copy()

            # Filter to relevant units
            if nome == "Taxa Ordinária" and isencao:
                df_per = df_per[~df_per["Unidade"].isin(sindicos)]
            elif df_src is df_e:
                df_per = df_per[df_per["Unidade"].isin(units_ext)]

            df_col = df_per[df_per[col].fillna(0) > 0.05].copy()
            n_pag  = len(df_col)

            # Step 4 — on-time / late
            no_v_df = df_col[df_col["dias_atraso"].fillna(999) <= carencia]
            atr_df  = df_col[df_col["dias_atraso"].fillna(-1)  >  carencia]
            n_no_v  = no_v_df["Unidade"].nunique()
            n_atr   = atr_df["Unidade"].nunique()

            # Step 4 — valor esperado: mean per on-time unit × n_expected
            if len(no_v_df) > 0:
                val_unit = float(no_v_df[col].mean())
            else:
                if nome == "Taxa Ordinária":
                    val_unit = float(params.get("taxa_ord_padrao") or 0)
                elif nome == "Medição e Leitura de Água":
                    val_unit = float(params.get("taxa_medicao") or 0)
                else:
                    te_m = next((te for te in params.get("taxas_extras", []) if te["nome"] == nome), None)
                    val_unit = float(te_m["valor"]) if te_m else 0.0
            val_esp = val_unit * n_esp

            # Valor base pago
            val_base = float(df_col[col].fillna(0).sum())

            # Multa e juros — pro-rata pela participação da coluna no total do boleto
            mj = 0.0
            for _, row_d in df_col.iterrows():
                rm = float(row_d.get("Receita com Multas") or 0)
                if rm <= 0:
                    continue
                row_base = sum(
                    float(row_d.get(vc) or 0)
                    for vc in VALOR_COLS_BASE if vc in df.columns
                )
                cv = float(row_d.get(col) or 0)
                mj += rm * (cv / row_base) if row_base > 0 else rm

            # Outros pagos (somente na linha de Taxa Ordinária; inclui ACORDO)
            outros = 0.0
            if add_o:
                if "Outros" in df.columns:
                    outros += float(df_per["Outros"].fillna(0).sum())
                df_ac_p = df_a[df_a["_mes"] == period]
                if len(df_ac_p):
                    outros += float(df_ac_p["Total"].fillna(0).sum())

            total_pago = val_base + mj + outros
            diff       = total_pago - val_esp
            paid_set   = set(df_col["Unidade"].unique())
            inad_set   = sorted(units_s - paid_set, key=lambda x: x.zfill(10))
            n_inad     = len(inad_set)
            lista_inad = "; ".join(inad_set) if inad_set else ""

            bg = LGRAY if idx % 2 == 0 else WHITE
            row_vals = [
                comp_str, n_pag, round(val_esp, 2),
                n_no_v, n_atr,
                round(val_base, 2), round(mj, 2), round(outros, 2),
                round(total_pago, 2), round(diff, 2), n_inad, lista_inad,
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=ci)
                c.value = v; c.fill = _f(bg); c.border = _bd()
                is_money = ci in (3, 6, 7, 8, 9, 10)
                is_diff  = (ci == 10)
                c.font = _ft(
                    size=10, bold=is_diff,
                    color=(C_RED if (is_diff and diff < 0) else
                           C_GRN if (is_diff and diff >= 0) else "1A2B3C"),
                )
                c.alignment = _al("right" if is_money else
                                  "left"  if ci == 12  else "center")
                if is_money:
                    c.number_format = '#,##0.00'
            ws.row_dimensions[r].height = 16; r += 1

            tot["pag"]  += n_pag;   tot["v_esp"] += val_esp; tot["no_v"] += n_no_v
            tot["atr"]  += n_atr;   tot["base"]  += val_base; tot["mj"]   += mj
            tot["out"]  += outros;  tot["tot"]   += total_pago; tot["inad"] += n_inad

        # Total row per payment section
        diff_t   = tot["tot"] - tot["v_esp"]
        tot_vals = [
            "Total", tot["pag"], round(tot["v_esp"], 2),
            tot["no_v"], tot["atr"],
            round(tot["base"], 2), round(tot["mj"], 2),
            round(tot["out"], 2),  round(tot["tot"], 2),
            round(diff_t, 2),      tot["inad"], "",
        ]
        for ci, v in enumerate(tot_vals, 1):
            c = ws.cell(row=r, column=ci)
            c.value = v; c.fill = _f(GOLD); c.border = _bd()
            is_money = ci in (3, 6, 7, 8, 9, 10)
            is_diff  = (ci == 10)
            c.font = _ft(
                bold=True, size=10,
                color=(C_RED if (is_diff and diff_t < 0) else
                       C_GRN if (is_diff and diff_t >= 0) else "1A2B3C"),
            )
            c.alignment = _al("right" if is_money else "center")
            if is_money:
                c.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 17; r += 2

    ws.freeze_panes = "A3"


def _gerar_excel(df, atrasados_sem_multa,
                 multa_inconsistente, multa_em_zero,
                 df_taxa, df_agua, df_medicao, df_extra, df_inadimplentes,
                 sem_param, sem_dados, campos_faltantes,
                 resultado, params, output_dir, session_id):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY="1A3C6E"; BLUE="1E88E5"; WHITE="FFFFFF"; LGRAY="F0F4F8"

    def hfill(c): return PatternFill("solid", fgColor=c)
    def hfont(bold=False, color=WHITE, size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def hbdr():
        s=Side(style="thin",color="B0BEC5"); return Border(left=s,right=s,top=s,bottom=s)
    def center(): return Alignment(horizontal="center",vertical="center")
    def left():   return Alignment(horizontal="left",  vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    def titulo(ws, texto, row, ncols=6):
        lc = get_column_letter(ncols)
        ws.merge_cells(f"A{row}:{lc}{row}")
        c=ws["A"+str(row)]; c.value=texto
        c.font=hfont(bold=True,size=11); c.fill=hfill(NAVY); c.alignment=center()
        ws.row_dimensions[row].height=22

    def kv(ws, row, label, valor, cor_val=NAVY):
        ws.merge_cells(f"A{row}:C{row}")
        cl=ws["A"+str(row)]; cl.value=label
        cl.font=hfont(bold=True,color=NAVY,size=10); cl.fill=hfill(LGRAY)
        cl.alignment=left(); cl.border=hbdr()
        ws.merge_cells(f"D{row}:F{row}")
        cv=ws["D"+str(row)]; cv.value=valor
        cv.font=hfont(bold=True,color=cor_val,size=10)
        cv.fill=hfill(WHITE); cv.alignment=left(); cv.border=hbdr()
        ws.row_dimensions[row].height=17

    def aba_df(wb, nome, df_aba, col_map, fill_hdr=BLUE, descricao=""):
        ncols = len(col_map)
        ws2 = wb.create_sheet(nome)
        if descricao:
            _add_nota(ws2, 1, ncols, descricao)
        hdr_row  = 2 if descricao else 1
        data_row = hdr_row + 1
        if df_aba is None or len(df_aba) == 0:
            ws2.cell(row=data_row, column=1).value = "Nenhuma ocorrência encontrada."
            ws2.cell(row=data_row, column=1).font  = hfont(bold=True, color="2E7D32", size=11)
            return
        colunas = list(col_map.keys())
        for ci, col in enumerate(colunas, 1):
            c = ws2.cell(row=hdr_row, column=ci)
            c.value = col_map[col]; c.font = hfont(bold=True, size=10)
            c.fill = hfill(fill_hdr); c.alignment = center(); c.border = hbdr()
            ws2.column_dimensions[get_column_letter(ci)].width = 20
        ws2.row_dimensions[hdr_row].height = 20
        for ri, (_, row) in enumerate(df_aba.iterrows(), data_row):
            rf = hfill(LGRAY) if ri % 2 == 0 else hfill(WHITE)
            for ci, col in enumerate(colunas, 1):
                c = ws2.cell(row=ri, column=ci)
                val = row.get(col, "") if col in row.index else ""
                if isinstance(val, float) and np.isnan(val): val = ""
                c.value = val; c.fill = rf; c.border = hbdr()
                c.font = hfont(color="1A2B3C", size=10); c.alignment = center()
            ws2.row_dimensions[ri].height = 16

    # Resumo
    ws["A1"].value="OGS Serviços — Análise e Conciliação de Cobranças"
    ws["A1"].font=hfont(bold=True,size=14); ws["A1"].fill=hfill(NAVY)
    ws.merge_cells("A1:F1"); ws["A1"].alignment=center(); ws.row_dimensions[1].height=30
    for col in ["A","B","C","D","E","F"]: ws.column_dimensions[col].width=18
    _add_nota(ws, 2, 6,
        "O QUE ESTA ABA VERIFICA: painel consolidado da conciliação entre a planilha de parâmetros "
        "e a base 004A (pagamentos realizados). Verificações realizadas: "
        "(1) Validação dos parâmetros — campos obrigatórios, unidades sem parâmetro e vice-versa; "
        "(2) Inadimplência — boletos NORMAL ausentes na 004A, excluindo síndico isento e unidades com ACORDO posterior; "
        "(3) Inconsistências de valores — taxa ordinária, água, medição e taxas extras; "
        "(4) Multa e atrasos — atrasados sem multa, multa fora da faixa esperada e multa sobre taxa zero. "
        "Para detalhes de cada verificação, consulte as abas específicas desta planilha."
    )

    r=3
    titulo(ws,"VISÃO GERAL",r); r+=1
    kv(ws,r,"Período",f"{resultado['periodo_inicio']} → {resultado['periodo_fim']}"); r+=1
    kv(ws,r,"Total emitido",resultado["total_emissao"]); r+=1
    kv(ws,r,"Registros / Unidades",f"{resultado['total_registros']} / {resultado['total_unidades']}"); r+=1

    r+=1; titulo(ws,"VALIDAÇÃO DOS PARÂMETROS",r); r+=1
    kv(ws,r,"Campos obrigatórios faltantes",
       ", ".join(campos_faltantes) if campos_faltantes else "Nenhum",
       "C62828" if campos_faltantes else "2E7D32"); r+=1
    kv(ws,r,"Unidades nos dados sem parâmetro",
       ", ".join(sem_param) if sem_param else "Nenhuma",
       "E65100" if sem_param else "2E7D32"); r+=1
    kv(ws,r,"Unidades no parâmetro sem dados",
       ", ".join(sem_dados) if sem_dados else "Nenhuma",
       "E65100" if sem_dados else "2E7D32"); r+=1

    r+=1; titulo(ws,"INADIMPLÊNCIA",r); r+=1
    kv(ws,r,"Boletos ausentes na 004A",resultado["qt_boletos_ausentes"],
       "C62828" if resultado["qt_boletos_ausentes"] else "2E7D32"); r+=1
    kv(ws,r,"Unidades inadimplentes",
       ", ".join(resultado["unidades_inadimplentes"]) if resultado["unidades_inadimplentes"] else "Nenhuma",
       "C62828" if resultado["unidades_inadimplentes"] else "2E7D32"); r+=1

    r+=1; titulo(ws,"INCONSISTÊNCIAS DE VALORES",r); r+=1
    kv(ws,r,"Taxa Ordinária divergente",resultado["qt_inconsist_taxa"],
       "C62828" if resultado["qt_inconsist_taxa"] else "2E7D32"); r+=1
    kv(ws,r,"Taxa de Água ausente/zero",resultado["qt_prob_agua"],
       "E65100" if resultado["qt_prob_agua"] else "2E7D32"); r+=1
    kv(ws,r,"Medição e Leitura divergente",resultado["qt_prob_medicao"],
       "E65100" if resultado["qt_prob_medicao"] else "2E7D32"); r+=1
    kv(ws,r,"Taxas Extras divergentes",resultado["qt_inconsist_extra"],
       "E65100" if resultado["qt_inconsist_extra"] else "2E7D32"); r+=1

    r+=1; titulo(ws,"MULTA E ATRASOS",r); r+=1
    kv(ws,r,"Atrasos reais (> carência)",resultado["atrasados_reais"]); r+=1
    kv(ws,r,"Atrasados reais sem multa",resultado["total_criticos"],
       "C62828" if resultado["total_criticos"] else "2E7D32"); r+=1
    kv(ws,r,"% multa médio / esperado",
       f"{resultado['multa_media']:.2f}% / {resultado['pct_multa_param']:.2f}%"); r+=1
    kv(ws,r,"Multa fora da faixa esperada",resultado["multa_inconsistente_qt"],
       "E65100" if resultado["multa_inconsistente_qt"] else "2E7D32"); r+=1
    kv(ws,r,"Multa sobre taxa ordinária zero",resultado["multa_em_zero_qt"],
       "E65100" if resultado["multa_em_zero_qt"] else "2E7D32"); r+=1

    # Fluxo de Pagamentos
    ws_fluxo = wb.create_sheet("Fluxo de Pagamentos")
    _aba_fluxo_pagamentos(ws_fluxo, df, params)

    # Inadimplência
    ws_inadin = wb.create_sheet("Inadimplência")
    _aba_inadimplencia(ws_inadin, df, params, df_inadimplentes)

    _car = resultado["carencia_param"]; _mul = resultado["pct_multa_param"]
    _med = resultado["ref_medicao"]

    # Abas de detalhe
    aba_df(wb,"Taxa Ordinária — Divergências", df_taxa,
           {"Unidade":"Unidade","Vencimento":"Vencimento",
            "Esperado (R$)":"Esperado (R$)","Encontrado (R$)":"Encontrado (R$)",
            "Diferença (R$)":"Diferença (R$)","Motivo":"Motivo"}, "C62828",
           descricao=(
               "O QUE ESTA ABA VERIFICA: registros NORMAL em que o maior valor de taxa ordinária "
               "cobrado no mês difere do valor esperado nos parâmetros em mais de R$ 0,05. "
               "CRITÉRIO: para unidades com múltiplos boletos NORMAL no mesmo mês (ex.: boleto de "
               "extras com taxa = 0), considera-se apenas o valor máximo do mês, evitando falsos positivos. "
               "Síndico isento: quando configurado nos parâmetros, o valor esperado é R$ 0,00 e qualquer "
               "cobrança de taxa ordinária é sinalizada como divergência."
           ))

    aba_df(wb,"Taxa de Água — Problemas", df_agua,
           {"Unidade":"Unidade","Vencimento":"Vencimento","Problema":"Problema","Valor":"Valor"}, "E65100",
           descricao=(
               "O QUE ESTA ABA VERIFICA: registros NORMAL com taxa de água ausente (NaN) ou igual a zero. "
               "CRITÉRIO: a taxa de água é cobrada mensalmente para todas as unidades com base no consumo "
               "individual; a ausência do valor no boleto NORMAL indica possível omissão operacional. "
               "Esta verificação é complementar à aba 'Inadimplência': enquanto aquela identifica boletos "
               "completamente ausentes, esta identifica boletos existentes com o campo água sem valor."
           ))

    aba_df(wb,"Medição — Divergências", df_medicao,
           {"Unidade":"Unidade","Vencimento":"Vencimento",
            "Problema":"Problema","Esperado":"Esperado (R$)","Encontrado":"Encontrado (R$)"}, "E65100",
           descricao=(
               f"O QUE ESTA ABA VERIFICA: registros NORMAL com valor de medição e leitura de água "
               f"ausente, zero ou divergente do valor de referência definido nos parâmetros. "
               f"CRITÉRIO: tolerância de R$ 0,05 em relação ao valor de referência ({_med}). "
               f"Cobranças ausentes/zero e cobranças com diferença acima da tolerância são listadas "
               f"com o motivo correspondente."
           ))

    aba_df(wb,"Taxas Extras — Divergências", df_extra,
           {"Taxa":"Taxa","Unidade":"Unidade","Vencimento":"Vencimento",
            "Esperado (R$)":"Esperado (R$)","Encontrado (R$)":"Encontrado (R$)",
            "Diferença (R$)":"Diferença (R$)"}, "E65100",
           descricao=(
               "O QUE ESTA ABA VERIFICA: registros EXTRA com valor cobrado divergente do esperado nos "
               "parâmetros (tolerância: R$ 0,05). CRITÉRIO: são verificadas apenas as unidades marcadas "
               "como 'com taxa extra' no cadastro de parâmetros. Dois tipos de ocorrência são sinalizados: "
               "(1) cobrança no período vigente com valor diferente do parâmetro; "
               "(2) cobrança fora do período vigente definido (início/fim) com valor acima de zero."
           ))

    aba_df(wb,"Atrasados sem Multa", atrasados_sem_multa,
           {"Unidade":"Unidade","Vencimento_dt":"Vencimento","Credito_dt":"Crédito",
            "dias_atraso":"Dias Atraso","Taxa Ordinária":"Taxa Ordinária (R$)",
            "Receita com Multas":"Multa (R$)"}, "C62828",
           descricao=(
               f"O QUE ESTA ABA VERIFICA: boletos NORMAL pagos com atraso real (crédito acima da "
               f"carência de {_car} dias após o vencimento) sem cobrança de multa registrada "
               f"(Receita com Multas = 0 ou ausente). "
               f"CRITÉRIO: a carência contempla a compensação bancária. A multa esperada é de {_mul} "
               f"sobre a taxa ordinária. A ausência de multa em pagamentos atrasados pode indicar "
               f"isenção não documentada ou falha operacional no sistema de cobrança."
           ))

    aba_df(wb,"Multa Inconsistente", multa_inconsistente,
           {"Unidade":"Unidade","Vencimento_dt":"Vencimento","dias_atraso":"Dias Atraso",
            "Taxa Ordinária":"Taxa Ord. (R$)","Receita com Multas":"Multa (R$)",
            "pct_multa_calc":"% Multa Calculado"}, "E65100",
           descricao=(
               f"O QUE ESTA ABA VERIFICA: boletos NORMAL com multa cobrada fora da faixa aceitável "
               f"em relação ao percentual configurado nos parâmetros ({_mul}). "
               f"CRITÉRIO: faixa aceitável de 75% a 150% do percentual esperado. Apenas boletos com "
               f"taxa ordinária > 0 são analisados para evitar distorções por divisão de valores zero. "
               f"O percentual calculado é: Receita com Multas ÷ Taxa Ordinária × 100."
           ))

    aba_df(wb,"Multa sobre Taxa Zero", multa_em_zero,
           {"Unidade":"Unidade","Vencimento_dt":"Vencimento",
            "Taxa Ordinária":"Taxa Ord. (R$)","Receita com Multas":"Multa (R$)"}, "E65100",
           descricao=(
               "O QUE ESTA ABA VERIFICA: boletos NORMAL com Receita com Multas > 0 e taxa ordinária "
               "igual a zero. CRITÉRIO: a multa deve incidir sobre o valor da taxa base; cobrar multa "
               "sobre taxa zero indica possível erro de lançamento no sistema de gestão — o valor de "
               "multa foi registrado em um boleto que não possui taxa ordinária associada."
           ))

    # Dados completos
    aba_df(wb,"Dados Completos", df,
           {"Unidade":"Unidade","Vencimento_dt":"Vencimento","Credito_dt":"Crédito",
            "dias_atraso":"Dias Atraso","Tipo Cobrança":"Tipo",
            "Taxa Ordinária":"Taxa Ord. (R$)","Taxa de Água":"Água (R$)",
            "Medição e Leitura de Água":"Medição (R$)",
            "Receita com Multas":"Multa (R$)","Total":"Total (R$)"}, NAVY,
           descricao=(
               "O QUE ESTA ABA CONTÉM: exportação integral da base 004A após normalização. "
               "Transformações aplicadas: datas de vencimento e crédito convertidas para datetime "
               "(formato DD/MM/AAAA); valores monetários convertidos de texto para numérico "
               "(suporte a formato brasileiro com ponto e vírgula); dias de atraso calculados como "
               "Crédito − Vencimento (negativo = pago antecipado; positivo = pago com atraso). "
               "Esta aba serve como fonte auditável para todas as demais análises desta planilha."
           ))

    out = os.path.join(output_dir, f"{session_id}_analise.xlsx")
    wb.save(out)
